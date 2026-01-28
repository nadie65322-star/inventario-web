from flask import Blueprint, render_template, request, jsonify, send_file
from inventario import db
from inventario.models import Item, Movimiento
from inventario.utils import normalize_text, semana_lunes_viernes
from datetime import datetime, date
import pandas as pd
import io

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from openpyxl.styles import PatternFill, Font, Alignment

bp = Blueprint('main', __name__)

# ==============================
# MESES EN ESPAÑOL
# ==============================
MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}

# ==============================
# INDEX
# ==============================
@bp.route('/')
def index():
    q = request.args.get('q', '')
    categoria = request.args.get('categoria', '')

    query = Item.query
    if q:
        query = query.filter(Item.nombre_normalizado.contains(normalize_text(q)))
    if categoria:
        query = query.filter_by(categoria=categoria)

    items = query.order_by(Item.nombre).all()
    return render_template('index.html', items=items, today=date.today())

# ==============================
# AÑADIR PRODUCTO
# ==============================
@bp.route('/add', methods=['POST'])
def add_item():
    fecha = request.form.get('fecha_vencimiento')

    item = Item(
        nombre=request.form.get('nombre'),
        nombre_normalizado=normalize_text(request.form.get('nombre')),
        categoria=request.form.get('categoria'),
        presentacion=request.form.get('presentacion'),
        lote=request.form.get('lote'),
        cantidad=int(request.form.get('cantidad') or 0),
        fecha_vencimiento=datetime.strptime(fecha, '%Y-%m-%d').date() if fecha else None
    )

    db.session.add(item)
    db.session.commit()

    db.session.add(Movimiento(
        item_id=item.id,
        tipo='inicial',
        cantidad=item.cantidad
    ))
    db.session.commit()

    return jsonify({'ok': True})

# ==============================
# INGRESO
# ==============================
@bp.route('/inc/<int:item_id>', methods=['POST'])
def inc(item_id):
    item = Item.query.get_or_404(item_id)
    cantidad = int(request.json['cantidad'])

    item.cantidad += cantidad
    db.session.add(Movimiento(
        item_id=item.id,
        tipo='ingreso',
        cantidad=cantidad
    ))
    db.session.commit()

    return jsonify({'cantidad': item.cantidad})

# ==============================
# RETIRO
# ==============================
@bp.route('/dec/<int:item_id>', methods=['POST'])
def dec(item_id):
    item = Item.query.get_or_404(item_id)
    cantidad = min(int(request.json['cantidad']), item.cantidad)

    item.cantidad -= cantidad
    db.session.add(Movimiento(
        item_id=item.id,
        tipo='retiro',
        cantidad=cantidad
    ))
    db.session.commit()

    return jsonify({'cantidad': item.cantidad})

# ==============================
# ELIMINAR PRODUCTO
# ==============================
@bp.route('/delete/<int:item_id>', methods=['POST'])
def delete_item(item_id):
    Movimiento.query.filter_by(item_id=item_id).delete()
    Item.query.filter_by(id=item_id).delete()
    db.session.commit()
    return jsonify({'ok': True})

# ==============================
# HISTORIAL
# ==============================
@bp.route('/historial/<int:item_id>')
def historial(item_id):
    movimientos = Movimiento.query.filter_by(item_id=item_id)\
        .order_by(Movimiento.fecha.desc()).all()

    return jsonify([
        {
            'fecha': m.fecha.strftime('%d/%m/%Y %H:%M'),
            'tipo': m.tipo,
            'cantidad': m.cantidad,
            'nota': m.nota
        } for m in movimientos
    ])

# ==============================
# EXCEL SEMANAL (3 HOJAS)
# ==============================
@bp.route('/export-excel-semana', methods=['POST'])
def export_excel_semana():
    lunes, viernes = semana_lunes_viernes()

    inventario = pd.read_sql("""
        SELECT nombre, categoria, presentacion, lote,
               fecha_vencimiento, cantidad
        FROM item
        ORDER BY nombre
    """, db.engine)

    historial = pd.read_sql("""
        SELECT i.nombre AS producto,
               m.fecha, m.tipo, m.cantidad
        FROM movimiento m
        JOIN item i ON i.id = m.item_id
        WHERE date(m.fecha) BETWEEN ? AND ?
        ORDER BY m.fecha
    """, db.engine, params=(lunes, viernes))

    movimientos = pd.read_sql("""
        SELECT date(m.fecha) AS fecha,
               m.tipo, m.cantidad
        FROM movimiento m
        WHERE date(m.fecha) BETWEEN ? AND ?
    """, db.engine, params=(lunes, viernes))

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        inventario.to_excel(writer, sheet_name='Inventario', index=False)
        historial.to_excel(writer, sheet_name='Historial semanal', index=False)

        wb = writer.book

        # ----- ESTILO ENCABEZADOS -----
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="D9D9D9")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        # ----- HOJA 3: RESUMEN SEMANAL -----
        ws = wb.create_sheet('Resumen semanal')

        fila = 1
        dias = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes']

        for i, dia in enumerate(dias):
            fecha = lunes + pd.Timedelta(days=i)
            data = movimientos[movimientos['fecha'] == fecha]

            ws.cell(row=fila, column=1, value=dia.upper()).font = Font(bold=True)
            fila += 1

            ws.append(['Ingresos', 'Retiros', 'Total'])

            ingreso = data[data['tipo'] == 'ingreso']['cantidad'].sum()
            retiro = data[data['tipo'] == 'retiro']['cantidad'].sum()

            ws.append([ingreso, retiro, ingreso - retiro])
            fila += 2

    output.seek(0)
    return send_file(output, download_name='inventario_semana.xlsx', as_attachment=True)

# ==============================
# PDF SEMANAL
# ==============================
@bp.route('/export-pdf')
def export_pdf():
    lunes, viernes = semana_lunes_viernes()
    mes = MESES_ES[viernes.month]

    items = Item.query.order_by(Item.nombre).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()
    elements = []

    titulo = f"Inventario del {lunes.day} al {viernes.day} de {mes} de {viernes.year}"
    elements.append(Paragraph(titulo, styles['Title']))
    elements.append(Paragraph("<br/>", styles['Normal']))

    tabla = [[
        'Nombre', 'Categoría', 'Presentación',
        'Lote', 'Vencimiento', 'Cantidad'
    ]]

    for i in items:
        tabla.append([
            i.nombre,
            i.categoria,
            i.presentacion or '',
            i.lote or '',
            i.fecha_vencimiento.strftime('%d/%m/%Y') if i.fecha_vencimiento else '',
            i.cantidad
        ])

    t = Table(tabla, repeatRows=1, colWidths=[160, 110, 120, 100, 110, 80])
    t.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN', (-1,1), (-1,-1), 'CENTER')
    ]))

    elements.append(t)
    doc.build(elements)
    buffer.seek(0)

    nombre = f"inventario_{lunes.day}_{viernes.day}_{mes}_{viernes.year}.pdf"

    return send_file(buffer, download_name=nombre, as_attachment=True)
